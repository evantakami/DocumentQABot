import gradio as gr
from langchain.text_splitter import MarkdownTextSplitter
from langchain.vectorstores import FAISS
from langchain.chains import RetrievalQA
from langchain.chat_models import ChatOpenAI
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain.schema import Document
from ollama_wrapper import OllamaChatWrapper

OLLAMA_HOST = "localhost"
OLLAMA_PORT = 11434
MODEL_NAME = 'elyza:jp8b'  # 指定されたモデル名

embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
llm = ChatOpenAI(model_name="gpt-4o-mini", temperature=0, api_key="")
ollama_chat = OllamaChatWrapper(base_url="http://localhost:11434", model_name="elyza:jp8b")

#Ragas(rag評価マトリックス)の導入
#Rerank Modelの導入
conversation_history = []
retriever = None
qa_chain = None

def process_uploaded_file(file):
    global retriever, qa_chain
    
    if file is None:
        return "ファイルがアップロードされていません。"
    
    file_path = file.name
    if not os.path.exists(file_path):
        return f"ファイルが見つかりません：{file_path}"
    
    with open(file_path, 'r', encoding='utf-8') as f:
        file_content = f.read()
    
    
    documents = [Document(page_content=file_content)]
    
    
    text_splitter = MarkdownTextSplitter()
    docs = text_splitter.split_documents(documents)
    
    print(f"分割後のチャンク数：{len(docs)}")
    
    vectorstore = FAISS.from_documents(docs, embeddings)
    retriever = vectorstore.as_retriever()
    
    qa_chain = RetrievalQA.from_chain_type(llm=llm, chain_type="stuff", retriever=retriever)
    
    return f"ドキュメントは {len(docs)} 個のチャンクに分割されました。"

def answer_question(question, use_online):
    global conversation_history, retriever, qa_chain
    
    if retriever is None:
        return "ファイルをアップロードしてください。", "", ""
    
    retrieved_docs = retriever.get_relevant_documents(question)
    context_text = "\n".join([doc.page_content for doc in retrieved_docs])
    
    history_text = "\n".join([f"Human: {q}\nAI: {a}" for q, a in conversation_history[-3:]])
    
    if use_online:
        prompt = f"Conversation history:\n{history_text}\n\nContext: {context_text}\n\nHuman: {question}\nAI:"
        answer = qa_chain.run(prompt)
    else:
        prompt = f"Conversation history:\n{history_text}\n\nContext: {context_text}\n\nHuman: {question}\nAI:"
        answer = ollama_chat.generate(prompt)
    
    conversation_history.append((question, answer))
    if len(conversation_history) > 10:
        conversation_history.pop(0)

    retrieved_chunks = "\n\n".join([f"Chunk {i+1}:\n{doc.page_content[:200]}..." for i, doc in enumerate(retrieved_docs)])
    
    return answer, retrieved_chunks, format_history()

def format_history():
    return "\n\n".join([f"Human: {q}\nAI: {a}" for q, a in conversation_history])

def create_interface():
    with gr.Blocks(css="""
    body {
        background-color: #f0f2f5;
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
    }
    .gr-button {
        padding: 10px 20px;
        font-size: 14px;
        background-color: #f57c00;
        color: white;
        border-radius: 5px;
    }
    """) as iface:
        gr.Markdown("# 📄 **DocumentQABot**")
        gr.Markdown("📁 **ファイルをアップロードして会話しましょう！💬**")
        
        with gr.Row():
            file_upload = gr.File(label="📄 **Markdownファイルをアップロード**")
        
        with gr.Row():
            process_btn = gr.Button("📥 **ファイルを処理**")
            process_output = gr.Textbox(label="✅ **処理結果**", scale=2)

        with gr.Row():
            with gr.Column(scale=2):
                question_input = gr.Textbox(lines=2, placeholder="❓ **問題を入力してください**")
                use_online = gr.Checkbox(label="🌐 **オンラインモデルを使用する (gpt-4o-mini)**", value=True)
                submit_btn = gr.Button("🚀 **問題を送信**")
                answer_output = gr.Textbox(lines=10, label="📝 **回答**")
            
            with gr.Column(scale=1):
                history_output = gr.Textbox(lines=20, label="📚 **対話履歴**", value="")
        
        retrieved_chunks_output = gr.Textbox(lines=5, label="🔍 **検索されたChunks**")
        
        process_btn.click(fn=process_uploaded_file, inputs=[file_upload], outputs=[process_output])
        submit_btn.click(fn=answer_question, inputs=[question_input, use_online], outputs=[answer_output, retrieved_chunks_output, history_output])


    return iface

if __name__ == "__main__":
    iface = create_interface()
    iface.launch(share=True)
import os
import logging
import pandas as pd
import tiktoken
import time
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from excel_to_markdown import process_excel_file
import gradio as gr
from langchain.chat_models import AzureChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.output_parsers import StructuredOutputParser
from langchain.schema import HumanMessage, ResponseSchema

# 配置日志
logging.basicConfig(
    filename="process.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

# 【修改】新增模型配置与费用参数
MODEL_CONFIG = {
    "GPT3.5": {
        "deployment_name": "gpt3.5-deployment",
        "input_cost": 0.000002,
        "output_cost": 0.000002
    },
    "4omini": {
        "deployment_name": "den-share-openai-gpt4o-mini",
        "input_cost": 0.000005,
        "output_cost": 0.000005
    }
}

# Azure OpenAI 配置（其它环境变量保持不变）
os.environ["OPENAI_API_TYPE"] = "azure"
os.environ["OPENAI_API_KEY"] = "your_api_key"
os.environ["OPENAI_API_BASE"] = "your_api_base"
os.environ["OPENAI_API_VERSION"] = "your_api_version"

# 计算 Token 数量
def count_tokens(text, model="gpt-4o-mini"):
    """
    指定したテキストの token 数を返します。
    """
    try:
        enc = tiktoken.encoding_for_model(model)
    except Exception:
        enc = tiktoken.get_encoding("cl100k_base")
    return len(enc.encode(text))

first_log_done = False
first_log_lock = threading.Lock()

# 定义模板
template_stage1 = """
あなたは文書チェックの専門家です。
以下のドキュメントをチェックし、問題があるかどうかを判断してください。
問題があれば改善提案も記載してください。
【チェック項目】
種別: {type_main}
種類: {type_sub}
確認内容: {check_item}
具体例/改善例: {example}
手順作成者コメント: {creator_comment}
【文書内容】
{document}
"""

prompt_stage1 = PromptTemplate(
    input_variables=["type_main", "type_sub", "check_item", "example", "creator_comment", "document"],
    template=template_stage1
)

response_schemas = [
    ResponseSchema(name="AI評価", description="評価結果。OK、NG、いずれか。"),
    ResponseSchema(name="改善案", description="NGの項目に対する修正提案。OKの場合は空白。")
]

output_parser = StructuredOutputParser.from_response_schemas(response_schemas)
format_instructions = output_parser.get_format_instructions()

template_stage2 = """
あなたはチェック項目を評価するアシスタントです。
以下のタスクのうち、1つの評価を実施してください。
[OK] or [NG]
{ai_comment}
"""

prompt_stage2 = PromptTemplate(
    input_variables=["ai_comment", "document"],
    partial_variables={"format_instructions": format_instructions},
    template=template_stage2
)

# 【修改】更新 process_check_item 函数，新增参数 llm_instance 与 selected_model，并计算 tokens 信息
def process_check_item(item, doc_text, llm_instance, selected_model):
    global first_log_done

    item_no = item.get("項番", "<no-id>")
    type_main = str(item.get("種別", ""))
    type_sub = str(item.get("種類", ""))
    check_item_text = str(item.get("確認内容", ""))
    example = str(item.get("実施例及び改善例", ""))
    creator_comment = str(item.get("手順作成者コメント", ""))

    try:
        comment_prompt = prompt_stage1.format(
            type_main=type_main,
            type_sub=type_sub,
            check_item=check_item_text,
            example=example,
            creator_comment=creator_comment,
            document=doc_text
        )

        ai_comment = llm_instance([HumanMessage(content=comment_prompt)]).content
        tokens_prompt1 = count_tokens(comment_prompt, model=MODEL_CONFIG[selected_model]["deployment_name"])
        tokens_response1 = count_tokens(ai_comment, model=MODEL_CONFIG[selected_model]["deployment_name"])

        with first_log_lock:
            if not first_log_done:
                logging.info("ユーザーのプロンプト内容: " + comment_prompt)
                logging.info("AIのコメント内容: " + ai_comment)
                first_log_done = True

    except Exception as e:
        logging.error(f"項番 {item_no}: ステージ1のコメント生成に失敗しました: {e}")
        return {
            "項番": item_no,
            "AI評価のコメント": f"エラー: {e}",
            "AI評価": "エラー",
            "改善案": f"エラー: {e}",
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0
        }

    try:
        result_prompt = prompt_stage2.format(ai_comment=ai_comment, document=doc_text)
        raw_output = llm_instance([HumanMessage(content=result_prompt)]).content
        result = output_parser.parse(raw_output)

        ai_hyouka = result.get("AI評価", "")
        kaizen = result.get("改善案", "")

        logging.info(f"項番 {item_no}: ステージ2の評価が正常に完了しました（評価: {ai_hyouka}）")
        tokens_prompt2 = count_tokens(result_prompt, model=MODEL_CONFIG[selected_model]["deployment_name"])
        tokens_response2 = count_tokens(raw_output, model=MODEL_CONFIG[selected_model]["deployment_name"])

    except Exception as e:
        logging.error(f"項番 {item_no}: ステージ2の評価生成に失敗しました: {e}")
        ai_hyouka = "エラー"
        kaizen = f"エラー: {e}"
        tokens_prompt2 = 0
        tokens_response2 = 0

    input_tokens = tokens_prompt1 + tokens_prompt2
    output_tokens = tokens_response1 + tokens_response2
    total_tokens = input_tokens + output_tokens

    return {
        "項番": item_no,
        "AI評価のコメント": ai_comment,
        "AI評価": ai_hyouka,
        "改善案": kaizen,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens
    }

# 【修改】更新 process_uploaded_file 函数，增加 selected_types, selected_model, llm_instance 参数  
# 并且：  
# 1. 读取 CSV 后不再过滤，而是在处理时判断是否符合 UI 勾选  
# 2. 输出到 Excel 时，按照 CSV 行号（第一行对应 Excel 第10行，第二行对应 Excel 第11行……）写入结果，未处理项在 AI評価のコメント 中写入“チェック対象外”
def process_uploaded_file(uploaded_excel_file, selected_types, selected_model, llm_instance):
    try:
        # 保存上传的 Excel 文件
        if isinstance(uploaded_excel_file, dict):
            file_name = uploaded_excel_file["name"]
            file_data = uploaded_excel_file["data"]
            os.makedirs("uploads", exist_ok=True)
            upload_path = os.path.join("uploads", file_name)
            with open(upload_path, "wb") as f:
                f.write(file_data)

        elif hasattr(uploaded_excel_file, "read"):
            file_name = os.path.basename(uploaded_excel_file.name)
            os.makedirs("uploads", exist_ok=True)
            upload_path = os.path.join("uploads", file_name)
            with open(upload_path, "wb") as f:
                f.write(uploaded_excel_file.read())
        else:
            file_name = os.path.basename(uploaded_excel_file)
            upload_path = uploaded_excel_file

        process_excel_file(upload_path)

        md_folder = os.path.join(os.path.dirname(upload_path), os.path.splitext(file_name)[0])
        if not os.path.exists(md_folder):
            return "Markdownファイルへの変換に失敗しました。フォルダが存在しません。", None

        # 【修改】读取 CSV，保持全部行，保证输出 Excel 时行号对应
        csv_file = "check_list.csv"
        df = pd.read_csv(csv_file, encoding="utf-8-sig")
        items = df.to_dict(orient="records")

        md_files = [f for f in os.listdir(md_folder) if f.lower().endswith(".md")]
        if not md_files:
            return "変換後のフォルダにMarkdownファイルがありません。", None

        # 为每个 md 文件生成结果，这里假设每个 md 文件生成一份结果（如果有多个，可根据需求合并）
        results_dict = {}
        cost_summary = {}
        for md_filename in md_files:
            md_path = os.path.join(md_folder, md_filename)
            with open(md_path, "r", encoding="utf-8") as f:
                doc_text = f.read()

            file_results = []
            total_input_tokens = 0
            total_output_tokens = 0
            # 【修改】按 CSV 顺序处理每个检查项
            for item in items:
                # 如果 UI 中选择了种别，且当前项的种别不在所选列表中，则直接标记为“チェック対象外”
                if selected_types and item.get("種別", "") not in selected_types:
                    result = {
                        "項番": item.get("項番", ""),
                        "AI評価のコメント": "チェック対象外",
                        "AI評価": "",
                        "改善案": "",
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "total_tokens": 0
                    }
                else:
                    result = process_check_item(item, doc_text, llm_instance, selected_model)
                file_results.append(result)
                total_input_tokens += result.get("input_tokens", 0)
                total_output_tokens += result.get("output_tokens", 0)
                time.sleep(2)  # API调用间隔

            results_dict[md_filename] = file_results

            model_cost = MODEL_CONFIG.get(selected_model, {"input_cost": 0, "output_cost": 0})
            cost_input = total_input_tokens * model_cost["input_cost"]
            cost_output = total_output_tokens * model_cost["output_cost"]
            total_cost = cost_input + cost_output
            cost_summary[md_filename] = (total_input_tokens, total_output_tokens, total_cost)

            # 【修改】将结果写入 Excel 模板  
            template_excel = "F-0168-2.xlsx"
            wb = openpyxl.load_workbook(template_excel)
            template_sheet_name = "商用作業手順書チェックリスト"
            if template_sheet_name not in wb.sheetnames:
                return f"テンプレートシート[{template_sheet_name}]が存在しません。", None
            template_sheet = wb[template_sheet_name]

            data_font = Font(name="Meiryo UI", size=10, bold=False)
            data_alignment = Alignment(horizontal="left", vertical="bottom")
            data_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            fields = ["AI評価のコメント", "AI評価", "改善案"]
            start_row = 10  # CSV第一行对应 Excel 第10行
            start_col = 11  # 从第11列开始写入

            # 遍历 file_results，与 CSV 行对应
            for idx, result in enumerate(file_results):
                excel_row = start_row + idx
                # 写入 AI評価のコメント
                cell = template_sheet.cell(row=excel_row, column=start_col)
                cell.value = result.get("AI評価のコメント", "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = data_fill

                # 写入 AI評価（假设在 start_col+1 列）
                cell = template_sheet.cell(row=excel_row, column=start_col+1)
                cell.value = result.get("AI評価", "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = data_fill

                # 写入 改善案（假设在 start_col+2 列）
                cell = template_sheet.cell(row=excel_row, column=start_col+2)
                cell.value = result.get("改善案", "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = data_fill

            # 输出 Excel 文件（此处每个 md 文件生成一个 Excel 文件，可根据需求合并）
            output_excel = f"チェック結果_{os.path.splitext(file_name)[0]}_{md_filename}.xlsx"
            wb.save(output_excel)
            # 这里仅记录第一个 md 文件的 cost_summary 用于展示
            break

        summary = "チェックが完了しました。\n"
        summary += f"Markdownファイル格納フォルダ: {md_folder}\n"
        for md_filename, cost in cost_summary.items():
            summary += f"{md_filename}: 入力tokens={cost[0]}, 出力tokens={cost[1]}, コスト=${cost[2]:.6f}\n"
        summary += f"作成されたExcelファイル: {output_excel}\n"

        return summary, output_excel

    except Exception as e:
        return f"処理中にエラーが発生しました: {e}", None

# 【修改】更新 run_interface 函数，新增模型与种别选择参数，并动态创建 llm_instance
def run_interface(uploaded_excel_file, selected_model, selected_types):
    try:
        # 根据所选模型创建 LLM 实例
        config = MODEL_CONFIG.get(selected_model)
        if config is None:
            raise ValueError("選択されたモデルが無効です。")
        llm_instance = AzureChatOpenAI(
            deployment_name=config["deployment_name"],
            openai_api_base=os.environ["OPENAI_API_BASE"],
            openai_api_version=os.environ["OPENAI_API_VERSION"],
            openai_api_key=os.environ["OPENAI_API_KEY"],
            temperature=0
        )
        results = process_uploaded_file(uploaded_excel_file, selected_types, selected_model, llm_instance)
        return results
    except Exception as e:
        return f"全体処理中にエラーが発生しました: {e}", None

# 【修改】读取 check_list.csv 中的种别选项，用于 UI 的 CheckboxGroup
try:
    df_check = pd.read_csv("check_list.csv", encoding="utf-8-sig")
    types_options = df_check["種別"].dropna().unique().tolist()
except Exception as e:
    logging.error(f"チェックリストの読み込みに失敗しました: {e}")
    types_options = []

# Gradio UI
with gr.Blocks() as demo:
    gr.Markdown("## AI 手順書チェックツール")
    with gr.Row():
        input_file = gr.File(label="Excel ファイルをアップロード", file_types=[".xlsx", ".xlsm"])
    with gr.Row():
        # 【修改】新增模型选择控件
        model_selection = gr.Radio(choices=["GPT3.5", "4omini"], label="モデル選択", value="4omini")
        # 【修改】新增种别过滤控件（空选表示全部）
        selected_types = gr.CheckboxGroup(choices=types_options, label="チェック項目の種別選択 (空欄なら全て対象)")
    with gr.Row():
        run_btn = gr.Button("処理開始")
    with gr.Row():
        output_message = gr.Textbox(label="処理情報", interactive=False, lines=10)
    with gr.Row():
        output_file = gr.File(label="結果の Excel ファイルをダウンロード", file_types=[".xlsx"])
    
    run_btn.click(fn=run_interface, inputs=[input_file, model_selection, selected_types], outputs=[output_message, output_file])

demo.launch()

